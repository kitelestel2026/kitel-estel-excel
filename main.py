from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
CORS(app)

STORES = [
    (1,'DANTE'),(2,'CORCEGA'),(3,'BORRELL'),(5,'VALENCIA'),
    (6,'CONSEJO'),(7,'BAILEN'),(9,'P.NOU'),(4,'DOMICILI'),
    (13,'MERCAT'),(14,'M.PLATJA'),(27,'MINETAS'),(33,'ST.CUGAT'),
    (28,'RAVAL'),(29,'PEDRA'),(23,'CALDES'),(32,'GLASS'),
    (10,'JULIA'),(16,'MES VI')
]

PINK='F4CCCC'; WHITE='FFFFFF'; PINK2='FBDADA'
BLUE_HDR='1F4E79'; BLUE_LT='BDD7EE'; YELLOW='FFF2CC'
COD_BG='D6E4F0'; COD_FG='1A7AAD'; RED_FG='C00000'

def brd():
    s=Side(style='thin',color='AAAAAA')
    return Border(left=s,right=s,top=s,bottom=s)

def st(cell,bg=WHITE,fg='000000',bold=False,align='center',size=9,wrap=False,italic=False):
    cell.font=Font(name='Arial',bold=bold,color=fg,size=size,italic=italic)
    cell.fill=PatternFill('solid',fgColor=bg)
    cell.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)
    cell.border=brd()

def fmt_date(iso):
    if not iso: return ''
    try:
        parts = str(iso).split('T')[0].split('-')
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except:
        return str(iso)

@app.route('/generate-excel',methods=['POST','OPTIONS'])
def generate_excel():
    if request.method=='OPTIONS':
        return '',200
    data=request.json
    orders=data.get('orders',{})
    date_str=data.get('date','')
    fecha_entrega=data.get('fecha_entrega', date_str)
    date_display = fmt_date(date_str)
    entrega_display = fmt_date(fecha_entrega)
    
    wb=openpyxl.Workbook()
    ws=wb.active
    ws.title='CUADRANTE'

    # ROW 1: store numbers
    for c,v in [(1,None),(2,None),(3,'F'),(4,None)]:
        ws.cell(1,c).value=v
    st(ws.cell(1,1),bg=BLUE_HDR,fg=WHITE,bold=True)
    st(ws.cell(1,2),bg=BLUE_HDR,fg=WHITE,bold=True)
    st(ws.cell(1,3),bg=YELLOW,bold=True)
    st(ws.cell(1,4),bg=BLUE_HDR,fg=WHITE,bold=True)
    col=5
    for i,(sid,snom) in enumerate(STORES):
        bg=PINK if i%2==0 else WHITE
        ws.cell(1,col).value=sid
        st(ws.cell(1,col),bg=bg,fg=BLUE_HDR,bold=True)
        st(ws.cell(1,col+1),bg=bg)
        ws.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+1)
        col+=2

    # ROW 2: dates + store names
    ws.cell(2,1).value='x'
    ws.cell(2,2).value=f"Pedido: {date_display} · Entrega: {entrega_display}"
    st(ws.cell(2,1),bg=BLUE_HDR,fg=WHITE,bold=True)
    st(ws.cell(2,2),bg=BLUE_LT,bold=True,size=9)
    st(ws.cell(2,3),bg=YELLOW,bold=True)
    st(ws.cell(2,4),bg=BLUE_LT,bold=True)
    col=5
    for i,(sid,snom) in enumerate(STORES):
        bg=PINK if i%2==0 else WHITE
        ws.cell(2,col).value=snom
        st(ws.cell(2,col),bg=bg,fg=BLUE_HDR,bold=True)
        st(ws.cell(2,col+1),bg=bg)
        ws.merge_cells(start_row=2,start_column=col,end_row=2,end_column=col+1)
        col+=2

    # ROW 3: headers
    for c,v,a in [(1,'COD.','center'),(2,'CONCEPTO','left'),(3,'F','center'),(4,'T','center')]:
        ws.cell(3,c).value=v
        st(ws.cell(3,c),bg=BLUE_HDR,fg=WHITE,bold=True,size=8,align=a)
    st(ws.cell(3,3),bg=YELLOW,fg='000000',bold=True,size=8)
    col=5
    for i,(sid,snom) in enumerate(STORES):
        bg=PINK if i%2==0 else WHITE
        ws.cell(3,col).value='S'; st(ws.cell(3,col),bg=bg,fg=RED_FG,bold=True,size=8)
        ws.cell(3,col+1).value='P'; st(ws.cell(3,col+1),bg=bg,fg=BLUE_HDR,bold=True,size=8)
        col+=2

    # DATA ROWS
    all_cods=set()
    for sid_str,sdata in orders.items():
        for l in sdata.get('lineas',[]):
            all_cods.add((l['cod'],l.get('concepto',''),l.get('tipo','')))
    sorted_cods=sorted(all_cods,key=lambda x:x[0])

    row=4
    for cod,concepto,tipo in sorted_cods:
        alt=(row%2==0)
        rbg='F5F5F5' if alt else WHITE
        total=0
        for sid_str,sdata in orders.items():
            for l in sdata.get('lineas',[]):
                if l['cod']==cod:
                    try: total+=float(l.get('pedido','0') or 0)
                    except: pass
        ws.cell(row,1).value=cod; st(ws.cell(row,1),bg=COD_BG,fg=COD_FG,bold=True)
        ws.cell(row,2).value=concepto; st(ws.cell(row,2),bg=rbg,align='left')
        ws.cell(row,3).value=total if total else ''; st(ws.cell(row,3),bg=YELLOW,bold=True)
        ws.cell(row,4).value=tipo; st(ws.cell(row,4),bg=rbg,fg='888888')
        col=5
        for i,(sid,snom) in enumerate(STORES):
            bg=(PINK if alt else PINK2) if i%2==0 else rbg
            s_val=''; p_val=''
            if str(sid) in orders:
                for l in orders[str(sid)].get('lineas',[]):
                    if l['cod']==cod:
                        s_val=l.get('sobra','') or ''
                        p_val=l.get('pedido','') or ''
                        break
            ws.cell(row,col).value=s_val
            st(ws.cell(row,col),bg=bg,fg=RED_FG if s_val else '000000',bold=bool(s_val))
            ws.cell(row,col+1).value=p_val
            st(ws.cell(row,col+1),bg=bg,fg=BLUE_HDR if p_val else '000000',bold=bool(p_val))
            col+=2
        ws.row_dimensions[row].height=14
        row+=1

    # ENCARGS
    has_enc=any(orders.get(str(sid),{}).get('encargs','') for sid,_ in STORES)
    if has_enc:
        ws.cell(row,2).value='ENCÀRRECS ESPECIALS'
        for c in [1,2,3,4]: st(ws.cell(row,c),bg=YELLOW,bold=True,align='left')
        col=5
        for i,(sid,snom) in enumerate(STORES):
            enc=orders.get(str(sid),{}).get('encargs','') or ''
            bg=PINK if i%2==0 else WHITE
            ws.cell(row,col).value=enc
            st(ws.cell(row,col),bg=bg,fg=BLUE_HDR,bold=bool(enc),align='left',wrap=True)
            st(ws.cell(row,col+1),bg=bg)
            col+=2
        ws.row_dimensions[row].height=30

    # Widths
    ws.column_dimensions['A'].width=7
    ws.column_dimensions['B'].width=24
    ws.column_dimensions['C'].width=5
    ws.column_dimensions['D'].width=4
    col=5
    for _ in STORES:
        ws.column_dimensions[get_column_letter(col)].width=6
        ws.column_dimensions[get_column_letter(col+1)].width=6
        col+=2

    ws.freeze_panes='E4'
    ws.row_dimensions[1].height=16
    ws.row_dimensions[2].height=18
    ws.row_dimensions[3].height=14

    buf=io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f'KITEL_entrega_{entrega_display.replace("/","-")}.xlsx'
    return send_file(buf,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,download_name=fname)

@app.route('/')
def health():
    return jsonify({'status':'ok'})

if __name__=='__main__':
    app.run(host='0.0.0.0',port=8080)
